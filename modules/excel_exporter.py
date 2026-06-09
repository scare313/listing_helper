"""
Excel Exporter – Generate marketplace-specific listing workbooks.

Creates multi-sheet Excel files with a master "All Products" sheet and
per-marketplace sheets (Amazon, Flipkart, Meesho).  Each sheet has
styled headers, auto-fitted columns, and properly spread bullet-point /
key-feature columns.

Variation rows are appended beneath their parent product row.
"""

from __future__ import annotations

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import settings

logger = logging.getLogger("listing_helper.modules.excel_exporter")

# ============================================================================
# Style Constants
# ============================================================================

HEADER_FILL_ALL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
HEADER_FILL_AMAZON = PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid")
HEADER_FILL_FLIPKART = PatternFill(start_color="2874F0", end_color="2874F0", fill_type="solid")
HEADER_FILL_MEESHO = PatternFill(start_color="570741", end_color="570741", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
DATA_FONT = Font(size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


# ============================================================================
# Helpers
# ============================================================================

def _safe_json(value: Any) -> list[str]:
    """Parse a JSON string (or list) into a list of strings.

    Returns an empty list on failure or if *value* is ``None``.
    """
    if value is None:
        return []
    if isinstance(value, list):
        return [str(v) for v in value]
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
            if isinstance(parsed, list):
                return [str(v) for v in parsed]
        except (json.JSONDecodeError, TypeError):
            pass
    return []


def _auto_fit_columns(ws) -> None:
    """Auto-fit column widths based on content, capped at 50 chars."""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
                max_length = max(max_length, min(len(val), 50))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(max_length + 2, 12)


def _apply_header_style(ws, header_fill: PatternFill) -> None:
    """Apply styling to the first (header) row of a worksheet."""
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _write_row(ws, row_num: int, values: list[Any]) -> None:
    """Write a list of values to a worksheet row with data styling."""
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        cell.alignment = WRAP_ALIGNMENT


# ============================================================================
# Sheet Builders
# ============================================================================

def _create_all_products_sheet(
    wb: Workbook,
    products: list[dict[str, Any]],
    variations_map: dict[int, list[dict[str, Any]]],
) -> None:
    """Create the master 'All Products' sheet with every column.

    Columns:
        SKU, Product Name, Brand, Category, Cost Price, Weight (g),
        Listing Status,
        Amazon Title, Amazon Bullet 1-5, Amazon Description,
        Amazon Search Terms, Amazon Price,
        Flipkart Title, Flipkart Key Feature 1-6, Flipkart Description,
        Flipkart Keywords, Flipkart Price,
        Meesho Title, Meesho Description, Meesho Price
    """
    ws = wb.active
    ws.title = "All Products"

    headers = [
        "SKU", "Product Name", "Brand", "Category", "Cost Price",
        "Weight (g)", "Listing Status",
        # Amazon (8 columns)
        "Amazon Title",
        "Amazon Bullet 1", "Amazon Bullet 2", "Amazon Bullet 3",
        "Amazon Bullet 4", "Amazon Bullet 5",
        "Amazon Description", "Amazon Search Terms", "Amazon Price",
        # Flipkart (10 columns)
        "Flipkart Title",
        "Flipkart Key Feature 1", "Flipkart Key Feature 2",
        "Flipkart Key Feature 3", "Flipkart Key Feature 4",
        "Flipkart Key Feature 5", "Flipkart Key Feature 6",
        "Flipkart Description", "Flipkart Keywords", "Flipkart Price",
        # Meesho (3 columns)
        "Meesho Title", "Meesho Description", "Meesho Price",
    ]
    ws.append(headers)
    _apply_header_style(ws, HEADER_FILL_ALL)

    row_num = 2
    for product in products:
        row = _build_all_row(product)
        _write_row(ws, row_num, row)
        row_num += 1

        # Append variation rows
        pid = product.get("id")
        for vc in variations_map.get(pid, []):
            var_row = _build_variation_all_row(product, vc)
            _write_row(ws, row_num, var_row)
            row_num += 1

    _auto_fit_columns(ws)


def _build_all_row(product: dict[str, Any]) -> list[Any]:
    """Build a single data row for the All Products sheet."""
    bullets = _safe_json(product.get("amazon_bullets"))
    features = _safe_json(product.get("flipkart_key_features"))

    return [
        product.get("sku"),
        product.get("name"),
        product.get("brand"),
        product.get("category"),
        product.get("cost_price"),
        product.get("weight_grams"),
        product.get("listing_status", "new"),
        # Amazon
        product.get("amazon_title"),
        *_pad_list(bullets, 5),
        product.get("amazon_description"),
        product.get("amazon_search_terms"),
        product.get("amazon_price"),
        # Flipkart
        product.get("flipkart_title"),
        *_pad_list(features, 6),
        product.get("flipkart_description"),
        product.get("flipkart_search_keywords"),
        product.get("flipkart_price"),
        # Meesho
        product.get("meesho_title"),
        product.get("meesho_description"),
        product.get("meesho_price"),
    ]


def _build_variation_all_row(
    parent: dict[str, Any],
    vc: dict[str, Any],
) -> list[Any]:
    """Build a variation row for the All Products sheet.

    Uses the variation SKU and variation-specific content where
    available, falling back to the parent product's data.
    """
    var_sku = vc.get("variation_sku") or vc.get("sku") or parent.get("sku")
    marketplace = vc.get("marketplace", "")
    var_bullets = _safe_json(vc.get("bullets"))

    # Start with parent data, override with variation content
    amazon_title = vc.get("title") if marketplace == "amazon" else parent.get("amazon_title")
    amazon_desc = vc.get("description") if marketplace == "amazon" else parent.get("amazon_description")
    amazon_terms = vc.get("keywords") if marketplace == "amazon" else parent.get("amazon_search_terms")

    flipkart_title = vc.get("title") if marketplace == "flipkart" else parent.get("flipkart_title")
    flipkart_desc = vc.get("description") if marketplace == "flipkart" else parent.get("flipkart_description")
    flipkart_kw = vc.get("keywords") if marketplace == "flipkart" else parent.get("flipkart_search_keywords")

    meesho_title = vc.get("title") if marketplace == "meesho" else parent.get("meesho_title")
    meesho_desc = vc.get("description") if marketplace == "meesho" else parent.get("meesho_description")

    # Bullets / features based on marketplace
    if marketplace == "amazon":
        a_bullets = var_bullets or _safe_json(parent.get("amazon_bullets"))
        f_features = _safe_json(parent.get("flipkart_key_features"))
    elif marketplace == "flipkart":
        a_bullets = _safe_json(parent.get("amazon_bullets"))
        f_features = var_bullets or _safe_json(parent.get("flipkart_key_features"))
    else:
        a_bullets = _safe_json(parent.get("amazon_bullets"))
        f_features = _safe_json(parent.get("flipkart_key_features"))

    var_label = f"{vc.get('variation_type', '')}: {vc.get('variation_value', '')}"

    return [
        var_sku,
        f"{parent.get('name', '')} ({var_label})",
        parent.get("brand"),
        parent.get("category"),
        parent.get("cost_price"),
        parent.get("weight_grams"),
        parent.get("listing_status", "new"),
        # Amazon
        amazon_title,
        *_pad_list(a_bullets, 5),
        amazon_desc,
        amazon_terms,
        parent.get("amazon_price"),
        # Flipkart
        flipkart_title,
        *_pad_list(f_features, 6),
        flipkart_desc,
        flipkart_kw,
        parent.get("flipkart_price"),
        # Meesho
        meesho_title,
        meesho_desc,
        parent.get("meesho_price"),
    ]


def _create_marketplace_sheet(
    wb: Workbook,
    marketplace_name: str,
    products: list[dict[str, Any]],
    variations_map: dict[int, list[dict[str, Any]]],
    header_fill: PatternFill,
) -> None:
    """Create a marketplace-specific sheet with only relevant columns.

    Args:
        wb: Target workbook.
        marketplace_name: Display name – 'Amazon', 'Flipkart', or 'Meesho'.
        products: List of product dicts.
        variations_map: product_id → list of variation content dicts.
        header_fill: Header background colour.
    """
    ws = wb.create_sheet(title=marketplace_name)
    mp_key = marketplace_name.lower()

    match mp_key:
        case "amazon":
            headers = [
                "SKU", "Title", "Brand",
                "Bullet 1", "Bullet 2", "Bullet 3", "Bullet 4", "Bullet 5",
                "Description", "Search Terms", "Price", "Category",
            ]
        case "flipkart":
            headers = [
                "SKU", "Title", "Brand",
                "Key Feature 1", "Key Feature 2", "Key Feature 3",
                "Key Feature 4", "Key Feature 5", "Key Feature 6",
                "Description", "Keywords", "Price", "Category",
            ]
        case "meesho":
            headers = [
                "SKU", "Title", "Brand",
                "Description", "Price", "Category",
            ]
        case _:
            headers = ["SKU", "Title", "Brand", "Description", "Price", "Category"]

    ws.append(headers)
    _apply_header_style(ws, header_fill)

    row_num = 2
    for product in products:
        row = _build_marketplace_row(mp_key, product)
        _write_row(ws, row_num, row)
        row_num += 1

        # Variation rows for this marketplace
        pid = product.get("id")
        for vc in variations_map.get(pid, []):
            if vc.get("marketplace") == mp_key:
                var_row = _build_variation_marketplace_row(mp_key, product, vc)
                _write_row(ws, row_num, var_row)
                row_num += 1

    _auto_fit_columns(ws)


def _build_marketplace_row(mp_key: str, product: dict[str, Any]) -> list[Any]:
    """Build a data row for a marketplace-specific sheet."""
    match mp_key:
        case "amazon":
            bullets = _safe_json(product.get("amazon_bullets"))
            return [
                product.get("sku"),
                product.get("amazon_title"),
                product.get("brand"),
                *_pad_list(bullets, 5),
                product.get("amazon_description"),
                product.get("amazon_search_terms"),
                product.get("amazon_price"),
                product.get("category"),
            ]
        case "flipkart":
            features = _safe_json(product.get("flipkart_key_features"))
            return [
                product.get("sku"),
                product.get("flipkart_title"),
                product.get("brand"),
                *_pad_list(features, 6),
                product.get("flipkart_description"),
                product.get("flipkart_search_keywords"),
                product.get("flipkart_price"),
                product.get("category"),
            ]
        case "meesho":
            return [
                product.get("sku"),
                product.get("meesho_title"),
                product.get("brand"),
                product.get("meesho_description"),
                product.get("meesho_price"),
                product.get("category"),
            ]
        case _:
            return [product.get("sku")]


def _build_variation_marketplace_row(
    mp_key: str,
    parent: dict[str, Any],
    vc: dict[str, Any],
) -> list[Any]:
    """Build a variation row for a marketplace-specific sheet."""
    var_sku = vc.get("variation_sku") or vc.get("sku") or parent.get("sku")
    var_bullets = _safe_json(vc.get("bullets"))

    match mp_key:
        case "amazon":
            bullets = var_bullets or _safe_json(parent.get("amazon_bullets"))
            return [
                var_sku,
                vc.get("title") or parent.get("amazon_title"),
                parent.get("brand"),
                *_pad_list(bullets, 5),
                vc.get("description") or parent.get("amazon_description"),
                vc.get("keywords") or parent.get("amazon_search_terms"),
                parent.get("amazon_price"),
                parent.get("category"),
            ]
        case "flipkart":
            features = var_bullets or _safe_json(parent.get("flipkart_key_features"))
            return [
                var_sku,
                vc.get("title") or parent.get("flipkart_title"),
                parent.get("brand"),
                *_pad_list(features, 6),
                vc.get("description") or parent.get("flipkart_description"),
                vc.get("keywords") or parent.get("flipkart_search_keywords"),
                parent.get("flipkart_price"),
                parent.get("category"),
            ]
        case "meesho":
            return [
                var_sku,
                vc.get("title") or parent.get("meesho_title"),
                parent.get("brand"),
                vc.get("description") or parent.get("meesho_description"),
                parent.get("meesho_price"),
                parent.get("category"),
            ]
        case _:
            return [var_sku]


def _pad_list(items: list[str], length: int) -> list[str | None]:
    """Pad a list to a fixed length with ``None`` values."""
    padded = list(items[:length])
    while len(padded) < length:
        padded.append(None)
    return padded


# ============================================================================
# Public API
# ============================================================================

def generate_export_excel(
    products: list[dict[str, Any]],
    variations_map: dict[int, list[dict[str, Any]]],
    filename: str | None = None,
) -> str:
    """Generate a multi-sheet Excel workbook with marketplace-specific sheets.

    Creates four sheets:
        1. **All Products** – master view with every column.
        2. **Amazon** – Amazon-specific columns only.
        3. **Flipkart** – Flipkart-specific columns only.
        4. **Meesho** – Meesho-specific columns only.

    Args:
        products: List of product dicts from the database.
        variations_map: Mapping of ``product_id`` → list of variation
            content dicts (from ``get_all_variation_content_for_product``).
        filename: Optional output filename. Auto-generated with a
            timestamp if not provided.

    Returns:
        Absolute path to the generated ``.xlsx`` file.
    """
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"listing_export_{timestamp}.xlsx"

    export_dir = Path(settings.EXPORT_PATH)
    export_dir.mkdir(parents=True, exist_ok=True)
    filepath = export_dir / filename

    wb = Workbook()

    # Sheet 1: All Products (master view)
    _create_all_products_sheet(wb, products, variations_map)

    # Sheet 2: Amazon
    _create_marketplace_sheet(wb, "Amazon", products, variations_map, HEADER_FILL_AMAZON)

    # Sheet 3: Flipkart
    _create_marketplace_sheet(wb, "Flipkart", products, variations_map, HEADER_FILL_FLIPKART)

    # Sheet 4: Meesho
    _create_marketplace_sheet(wb, "Meesho", products, variations_map, HEADER_FILL_MEESHO)

    wb.save(str(filepath))
    logger.info("Export saved to: %s (%d products)", filepath, len(products))
    return str(filepath.resolve())
